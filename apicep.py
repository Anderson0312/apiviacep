import openpyxl
import requests

bairro = 'Rio%20de%20janeiro'

# Função para consultar a API dos Correios
def consulta_cep(endereco):
    try:
        url = f'https://viacep.com.br/ws/RJ/{bairro}/{endereco}/json/'
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            if isinstance(data, list) and len(data) == 0:
                print(f"CEP não encontrado para o endereço '{endereco}'")
                return None
            else:
                return data[0].get('cep')
        else:
            print(f"Erro ao consultar CEP para o endereço '{endereco}': Código {response.status_code}")
            return None
    except Exception as e:
        print(f"Erro ao consultar CEP para o endereço '{endereco}': {str(e)}")
        return None


# Abre o arquivo Excel
nome_arquivo = 'seu_arquivo.xlsx'  # Substitua 'seu_arquivo.xlsx' pelo nome do seu arquivo
workbook = openpyxl.load_workbook(nome_arquivo)
sheet = workbook.active

total_linhas = sheet.max_row - 1  # Ignora o cabeçalho
linhas_processadas = 0

# Loop pelas linhas do arquivo
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    endereco = row[0].value
    cep = consulta_cep(endereco)
    if cep:
        # Escreve o CEP na próxima coluna
        sheet.cell(row=row[0].row, column=2, value=cep)

    linhas_processadas += 1
    print(f'{linhas_processadas}/{total_linhas} linhas processadas.', end='\r')

# Salva as alterações no arquivo Excel
workbook.save(nome_arquivo)
print('\nConcluído! As informações foram atualizadas no arquivo Excel.')
